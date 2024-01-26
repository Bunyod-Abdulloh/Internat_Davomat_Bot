from aiogram import types
from openpyxl.styles import Side, Border, Alignment
from openpyxl.workbook import Workbook

from keyboards.inline.admin_edit_educator_buttons import edit_educators
from loader import db


# admin educators line 26
async def educator_main_first(educator_id: str, call: types.CallbackQuery, educator: list, employee: str,
                              first_phone: str, class_: str):
    await call.message.edit_text(
        text=f"{employee}: {educator[1]}"
             f"\n{first_phone}: {educator[2]}"
             f"\n{class_}: {educator[4]}",
        reply_markup=await edit_educators(
            telegram_id=educator_id
        )
    )


# admin educators line 31
async def educator_main_second(educator_id: str, call: types.CallbackQuery, educator: list, employee: str,
                               first_phone: str, second_phone: str, class_: str):
    await call.message.edit_text(
        text=f"{employee}: {educator[1]}"
             f"\n{first_phone}: {educator[2]}"
             f"\n{second_phone}: {educator[3]}"
             f"\n{class_}: {educator[4]}",
        reply_markup=await edit_educators(
            telegram_id=educator_id,
            second_phone=True
        )
    )


async def send_xlsx(current_date, all_levels, new_date):
    wb = Workbook()
    ws = wb.active
    ws.append(['Sana:', new_date, '', '', '', ''])
    ws.append(['№', 'Sinf', 'Umumiy o\'quvchilar soni', 'Kelgan o\'quvchilar soni',
               'Kelmagan o\'quvchilar soni', 'Yotoqxonada qoladigan o\'quvchilar soni'])
    one_four_all = await db.morning_report_levels(
        first_value=1, second_value=4, current_date=current_date
    )
    one_four_present = await db.morning_report(
        first_value=1, second_value=4, check_teacher='✅', current_date=current_date
    )
    one_four_absent = await db.morning_report(
        first_value=1, second_value=4, check_teacher='☑️', current_date=current_date
    )
    one_four_explicable = await db.morning_report(
        first_value=1, second_value=4, check_teacher='❎', current_date=current_date
    )
    one_four_abs_exp = one_four_absent + one_four_explicable
    five_seven_all = await db.morning_report_levels(
        first_value=5, second_value=7, current_date=current_date
    )
    five_seven_present = await db.morning_report(
        first_value=5, second_value=7, check_teacher='✅', current_date=current_date
    )
    five_seven_absent = await db.morning_report(
        first_value=5, second_value=7, check_teacher='☑️', current_date=current_date
    )
    five_seven_explicable = await db.morning_report(
        first_value=5, second_value=7, check_teacher='❎', current_date=current_date
    )
    five_seven_abs_exp = five_seven_absent + five_seven_explicable
    one_seven_all = await db.morning_report_levels(
        first_value=1, second_value=7, current_date=current_date
    )
    one_seven_present = await db.morning_report(
        first_value=1, second_value=7, check_teacher='✅', current_date=current_date
    )
    one_seven_absent = await db.morning_report(
        first_value=1, second_value=7, check_teacher='☑️', current_date=current_date
    )
    one_seven_explicable = await db.morning_report(
        first_value=1, second_value=7, check_teacher='❎', current_date=current_date
    )
    one_seven_abs_exp = one_seven_absent + one_seven_explicable
    eight_eleven_all = await db.morning_report_levels(
        first_value=8, second_value=11, current_date=current_date
    )
    eight_eleven_present = await db.morning_report(
        first_value=8, second_value=11, check_teacher='✅', current_date=current_date
    )
    eight_eleven_absent = await db.morning_report(
        first_value=8, second_value=11, check_teacher='☑️', current_date=current_date
    )
    eight_eleven_explicable = await db.morning_report(
        first_value=8, second_value=11, check_teacher='❎', current_date=current_date
    )
    eight_eleven_abs_exp = eight_eleven_absent + eight_eleven_explicable
    one_eleven_all = await db.morning_report_levels(
        first_value=1, second_value=11, current_date=current_date
    )
    one_eleven_present = await db.morning_report(
        first_value=1, second_value=11, check_teacher='✅', current_date=current_date
    )
    one_eleven_absent = await db.morning_report(
        first_value=1, second_value=11, check_teacher='☑️', current_date=current_date
    )
    one_eleven_explicable = await db.morning_report(
        first_value=1, second_value=11, check_teacher='❎', current_date=current_date
    )
    one_eleven_abs_exp = one_eleven_absent + one_eleven_explicable
    count = 0
    for n in all_levels:
        level = f'{n[0]}-{n[1]}'
        all_students = await db.count_everyday_attendance(current_date=current_date, level=level)
        present = await db.count_everyday_morning_check(current_date=current_date, level=level, check_teacher='✅')
        absent = await db.count_everyday_morning_check(current_date=current_date, level=level, check_teacher='☑️')
        explicable = await db.count_everyday_morning_check(current_date=current_date, level=level,
                                                           check_teacher='❎')
        abs_exp = absent + explicable
        count += 1
        ws.append([count, level, all_students, present, abs_exp, ''])
        if level == '4-V':
            ws.append(['', '1-4 sinflar', one_four_all, one_four_present, one_four_abs_exp, ''])
        elif level == '7-V':
            ws.append(['', '5-7 sinflar', five_seven_all, five_seven_present, five_seven_abs_exp, ''])
            ws.append(['', '1-7 sinflar', one_seven_all, one_seven_present, one_seven_abs_exp, ''])
        elif level == '11-A':
            ws.append(['', '8-11 sinflar', eight_eleven_all, eight_eleven_present, eight_eleven_abs_exp, ''])
            ws.append(['', '1-11 sinflar', one_eleven_all, one_eleven_present, one_eleven_abs_exp, ''])
    max_row = ws.max_row
    start_cell = ws['A2']
    end_cell = ws[f'F{max_row}']
    border_style = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))
    alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=start_cell.row, max_row=end_cell.row,
                            min_col=start_cell.column, max_col=end_cell.column):
        for cell in row:
            cell.border = border_style
            cell.alignment = alignment_style
    wb.save("Kunlik_hisobot.xlsx")

    line = '\n------------------------------------\n'
    empty = 'ㅤ'
    double_empty = empty + empty
    text = (f'Joriy sana: {new_date}\n\n<b>1-4 sinflar:</b>'
            f'\n{empty}📌Jami: {one_four_all}'
            f'\n{double_empty}✅ Kelganlar: {one_four_present}'
            f'\n{double_empty}❌ Kelmaganlar: {one_four_abs_exp}'
            f'{line}'
            f'<b>5-7 sinflar:</b>'
            f'\n{empty}📌Jami: {five_seven_all}'
            f'\n{double_empty}✅ Kelganlar: {five_seven_present}'
            f'\n{double_empty}❌ Kelmaganlar: {five_seven_abs_exp}'
            f'\n\n<b>1-7 sinflar:</b>'
            f'\n{empty}📌Jami: {one_seven_all}'
            f'\n{double_empty}✅ Kelganlar: {one_seven_present}'
            f'\n{double_empty}❌ Kelmaganlar: {one_seven_abs_exp}'
            f'{line}'
            f'<b>8-11 sinflar:</b>'
            f'\n{empty}📌Jami: {five_seven_all}'
            f'\n{double_empty}✅ Kelganlar: {five_seven_present}'
            f'\n{double_empty}❌ Kelmaganlar: {five_seven_abs_exp}'
            f'\n\n<b>1-11 sinflar:</b>'
            f'\n{empty}📌Jami: {one_eleven_all}'
            f'\n{double_empty}✅ Kelganlar: {one_eleven_present}'
            f'\n{double_empty}❌ Kelmaganlar: {one_eleven_abs_exp}'
            )
    return text
