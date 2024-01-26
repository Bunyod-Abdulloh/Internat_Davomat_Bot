import os
from datetime import datetime

from aiogram import types
from magic_filter import F
from openpyxl.styles import Border, Side

from data.config import ADMINS
from keyboards.default.admin_custom_buttons import admin_attendance_cbuttons
from openpyxl import Workbook
from loader import dp, db, bot


@dp.message_handler(F.text == 'Davomat', state='*')
async def view_attendance_main(message: types. Message):
    await message.answer(
        text='Kerakli bo\'limni tanlang:', reply_markup=admin_attendance_cbuttons
    )


@dp.message_handler(F.text == 'Sinflar davomati', state='*')
async def view_levels_attendance(message: types.Message):
    current_date = datetime.now().date()
    all_levels = await db.select_all_classes()
    date = f'Sana: {current_date}\n\n'
    submitted = 'Davomat topshirgan sinflar:\n'
    not_submitted = '\nDavomat topshirmagan sinflar:\n'
    count_yes = 0
    count_no = 0
    for level in all_levels:
        level_ = f'{level[0]}-{level[1]}'
        check_level = await db.select_check_work_teacher(checked_date=current_date, level=level_)
        if check_level is None:
            count_no += 1
            not_submitted += f'{count_no}. {level_}\n'
        else:
            count_yes += 1
            submitted += f'{count_yes}. {level_}\n'
    await message.answer(
        text=date + submitted + not_submitted
    )


@dp.message_handler(F.text == 'Kunlik davomat', state='*')
async def view_everydays_attendance(message: types.Message):
    current_date = datetime.now().date()
    all_levels = await db.all_levels_attendance(current_date=current_date)
    all_students = str()
    present = str()
    absent = str()
    explicable = str()
    dormitory = str()
    one_four_all = await db.morning_report_levels(
        first_value=1, second_value=4, current_date=current_date
    )
    one_four_present = await db.morning_report(
        first_value=1, second_value=4, check_teacher='✅', current_date=current_date
    )
    one_four_absent = await db.morning_report(
        first_value=1, second_value=4, check_teacher='☑️', current_date=current_date
    )
    one_four_explicable = await db.morning_report(
        first_value=1, second_value=4, check_teacher='❎', current_date=current_date
    )
    one_four_abs_exp = one_four_absent + one_four_explicable
    five_seven_all = await db.morning_report_levels(
        first_value=5, second_value=7, current_date=current_date
    )
    five_seven_present = await db.morning_report(
        first_value=5, second_value=7, check_teacher='✅', current_date=current_date
    )
    five_seven_absent = await db.morning_report(
        first_value=5, second_value=7, check_teacher='☑️', current_date=current_date
    )
    five_seven_explicable = await db.morning_report(
        first_value=5, second_value=7, check_teacher='❎', current_date=current_date
    )
    five_seven_abs_exp = five_seven_absent + five_seven_explicable
    one_seven_all = await db.morning_report_levels(
        first_value=1, second_value=7, current_date=current_date
    )
    one_seven_present = await db.morning_report(
        first_value=1, second_value=7, check_teacher='✅', current_date=current_date
    )
    one_seven_absent = await db.morning_report(
        first_value=1, second_value=7, check_teacher='☑️', current_date=current_date
    )
    one_seven_explicable = await db.morning_report(
        first_value=1, second_value=7, check_teacher='❎', current_date=current_date
    )
    one_seven_abs_exp = one_seven_absent + one_seven_explicable
    eight_eleven_all = await db.morning_report_levels(
        first_value=8, second_value=11, current_date=current_date
    )
    eight_eleven_present = await db.morning_report(
        first_value=8, second_value=11, check_teacher='✅', current_date=current_date
    )
    eight_eleven_absent = await db.morning_report(
        first_value=8, second_value=11, check_teacher='☑️', current_date=current_date
    )
    eight_eleven_explicable = await db.morning_report(
        first_value=8, second_value=11, check_teacher='❎', current_date=current_date
    )
    eight_eleven_abs_exp = eight_eleven_absent + eight_eleven_explicable
    one_eleven_all = await db.morning_report_levels(
        first_value=1, second_value=11, current_date=current_date
    )
    one_eleven_present = await db.morning_report(
        first_value=1, second_value=11, check_teacher='✅', current_date=current_date
    )
    one_eleven_absent = await db.morning_report(
        first_value=1, second_value=11, check_teacher='☑️', current_date=current_date
    )
    one_eleven_explicable = await db.morning_report(
        first_value=1, second_value=11, check_teacher='❎', current_date=current_date
    )
    one_eleven_abs_exp = one_eleven_absent + one_eleven_explicable
    wb = Workbook()
    ws = wb.active
    border_style = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    ws.append(['Sana:', current_date, '', '', '', ''])
    ws.append(['№', 'Sinf', 'Umumiy o\'quvchilar soni', 'Kelgan o\'quvchilar soni',
               'Kelmagan o\'quvchilar soni', 'Yotoqxonada qoladigan o\'quvchilar soni'])
    count = 0
    for n in all_levels:
        level = f'{n[0]}-{n[1]}'
        all_students = await db.count_everyday_attendance(current_date=current_date, level=level)
        present = await db.count_everyday_morning_check(current_date=current_date, level=level, check_teacher='✅')
        absent = await db.count_everyday_morning_check(current_date=current_date, level=level, check_teacher='☑️')
        explicable = await db.count_everyday_morning_check(current_date=current_date, level=level, check_teacher='❎')
        abs_exp = absent + explicable
        count += 1
        ws.append([count, level, all_students, present, abs_exp, ''])
        if level == '4-V':
            ws.append(['', '1-4 sinflar', one_four_all, one_four_present, one_four_abs_exp, ''])
        elif level == '7-V':
            ws.append(['', '5-7 sinflar', five_seven_all, five_seven_present, five_seven_abs_exp, ''])
            ws.append(['', '1-7 sinflar', one_seven_all, one_seven_present, one_seven_abs_exp, ''])
        elif level == '11-A':
            ws.append(['', '8-11 sinflar', eight_eleven_all, eight_eleven_present, eight_eleven_abs_exp, ''])
            ws.append(['', '1-11 sinflar', one_eleven_all, one_eleven_present, one_eleven_abs_exp, ''])
    second_row = 2
    max_column = ws.max_column
    last_row = ws.max_row
    cell = ws[second_row]
    print(second_row, max_column, last_row)
    cell.border = border_style
    wb.save("Kunlik_hisobot.xlsx")

    await bot.send_document(chat_id=ADMINS[0],
                            document=types.InputFile(path_or_bytesio="Kunlik_hisobot.xlsx")
                            )
    os.remove("Kunlik_hisobot.xlsx")


@dp.message_handler(F.text == 'Haftalik davomat', state='*')
async def view_weekly_attendance(message: types.Message):
    pass


@dp.message_handler(F.text == 'Oylik davomat', state='*')
async def view_monthly_attendance(message: types.Message):
    pass
